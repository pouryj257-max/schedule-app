"""일정관리 웹 애플리케이션 - Flask Backend"""
import os
import json
import sqlite3
import hashlib
from datetime import datetime, timedelta
from functools import wraps
from flask import (
    Flask, render_template, request, jsonify, redirect,
    url_for, session, flash, send_file, g
)
import openpyxl

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'schedule-app-secret-key-2026')
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
DB_PATH = os.path.join(os.path.dirname(__file__), 'schedule.db')

# ── DB helpers ──────────────────────────────────────────────

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db:
        db.close()

def init_db():
    db = sqlite3.connect(DB_PATH)
    db.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            password TEXT NOT NULL,
            team TEXT DEFAULT '',
            role TEXT DEFAULT 'user',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT NOT NULL,
            date TEXT NOT NULL,
            shift_type TEXT DEFAULT '',
            detail TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(employee_id, date)
        );
        CREATE TABLE IF NOT EXISTS memos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT NOT NULL,
            date TEXT NOT NULL,
            content TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_schedules_emp_date ON schedules(employee_id, date);
        CREATE INDEX IF NOT EXISTS idx_memos_emp_date ON memos(employee_id, date);
    ''')
    # 관리자 계정 생성
    try:
        db.execute(
            "INSERT OR IGNORE INTO users (employee_id, name, password, role) VALUES (?, ?, ?, ?)",
            ('admin', '관리자', hash_pw('admin'), 'admin')
        )
        db.commit()
    except Exception:
        pass
    db.close()

def hash_pw(pw):
    return hashlib.sha256(pw.encode('utf-8')).hexdigest()

# ── Auth ────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': '로그인이 필요합니다'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': '관리자 권한이 필요합니다'}), 403
            flash('관리자 권한이 필요합니다', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

# ── Routes: Auth ────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        emp_id = request.form.get('employee_id', '').strip()
        password = request.form.get('password', '').strip()
        db = get_db()
        user = db.execute(
            "SELECT * FROM users WHERE employee_id = ? AND password = ?",
            (emp_id, hash_pw(password))
        ).fetchone()
        if user:
            session['user_id'] = user['employee_id']
            session['user_name'] = user['name']
            session['role'] = user['role']
            return redirect(url_for('index'))
        flash('사번 또는 비밀번호가 올바르지 않습니다', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── Routes: Main ────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    return render_template('index.html')

# ── API: Schedules ──────────────────────────────────────────

@app.route('/api/schedules')
@login_required
def api_get_schedules():
    """달력용 일정 조회. ?year=2026&month=3 또는 ?start=2026-03-01&end=2026-03-31"""
    db = get_db()
    emp_id = session['user_id']
    is_admin = session.get('role') == 'admin'
    target_emp = request.args.get('employee_id', emp_id if not is_admin else None)

    year = request.args.get('year')
    month = request.args.get('month')
    start = request.args.get('start')
    end = request.args.get('end')

    if year and month:
        start = f"{year}-{int(month):02d}-01"
        if int(month) == 12:
            end = f"{int(year)+1}-01-01"
        else:
            end = f"{year}-{int(month)+1:02d}-01"

    query = "SELECT * FROM schedules WHERE date >= ? AND date < ?"
    params = [start, end]
    if target_emp:
        query += " AND employee_id = ?"
        params.append(target_emp)
    elif not is_admin:
        query += " AND employee_id = ?"
        params.append(emp_id)

    rows = db.execute(query, params).fetchall()
    results = []
    for r in rows:
        results.append({
            'id': r['id'],
            'employee_id': r['employee_id'],
            'date': r['date'],
            'shift_type': r['shift_type'],
            'detail': r['detail'],
        })
    return jsonify(results)

@app.route('/api/schedules', methods=['POST'])
@login_required
def api_upsert_schedule():
    data = request.json
    db = get_db()
    emp_id = data.get('employee_id', session['user_id'])
    if session.get('role') != 'admin' and emp_id != session['user_id']:
        return jsonify({'error': '권한 없음'}), 403

    db.execute('''
        INSERT INTO schedules (employee_id, date, shift_type, detail, updated_at)
        VALUES (?, ?, ?, ?, datetime('now'))
        ON CONFLICT(employee_id, date) DO UPDATE SET
            shift_type=excluded.shift_type,
            detail=excluded.detail,
            updated_at=datetime('now')
    ''', (emp_id, data['date'], data.get('shift_type', ''), data.get('detail', '')))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/schedules/<int:sid>', methods=['DELETE'])
@login_required
def api_delete_schedule(sid):
    db = get_db()
    row = db.execute("SELECT * FROM schedules WHERE id = ?", (sid,)).fetchone()
    if not row:
        return jsonify({'error': '일정 없음'}), 404
    if session.get('role') != 'admin' and row['employee_id'] != session['user_id']:
        return jsonify({'error': '권한 없음'}), 403
    db.execute("DELETE FROM schedules WHERE id = ?", (sid,))
    db.commit()
    return jsonify({'success': True})

# ── API: Memos ──────────────────────────────────────────────

@app.route('/api/memos')
@login_required
def api_get_memos():
    db = get_db()
    emp_id = request.args.get('employee_id', session['user_id'])
    date = request.args.get('date')
    start = request.args.get('start')
    end = request.args.get('end')

    if date:
        rows = db.execute(
            "SELECT * FROM memos WHERE employee_id = ? AND date = ?",
            (emp_id, date)
        ).fetchall()
    elif start and end:
        rows = db.execute(
            "SELECT * FROM memos WHERE employee_id = ? AND date >= ? AND date < ?",
            (emp_id, start, end)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM memos WHERE employee_id = ? ORDER BY date DESC LIMIT 100",
            (emp_id,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/memos', methods=['POST'])
@login_required
def api_save_memo():
    data = request.json
    db = get_db()
    emp_id = data.get('employee_id', session['user_id'])
    memo_id = data.get('id')

    if memo_id:
        db.execute(
            "UPDATE memos SET content=?, updated_at=datetime('now') WHERE id=? AND employee_id=?",
            (data['content'], memo_id, emp_id)
        )
    else:
        db.execute(
            "INSERT INTO memos (employee_id, date, content) VALUES (?, ?, ?)",
            (emp_id, data['date'], data['content'])
        )
    db.commit()
    return jsonify({'success': True})

@app.route('/api/memos/<int:mid>', methods=['DELETE'])
@login_required
def api_delete_memo(mid):
    db = get_db()
    db.execute("DELETE FROM memos WHERE id = ? AND employee_id = ?",
               (mid, session['user_id']))
    db.commit()
    return jsonify({'success': True})

# ── API: Search ─────────────────────────────────────────────

@app.route('/api/search')
@login_required
def api_search():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    db = get_db()
    emp_id = session['user_id']
    is_admin = session.get('role') == 'admin'

    results = []
    # 일정 검색
    if is_admin:
        rows = db.execute(
            "SELECT s.*, u.name FROM schedules s LEFT JOIN users u ON s.employee_id = u.employee_id "
            "WHERE s.shift_type LIKE ? OR s.detail LIKE ? ORDER BY s.date DESC LIMIT 50",
            (f'%{q}%', f'%{q}%')
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT s.*, u.name FROM schedules s LEFT JOIN users u ON s.employee_id = u.employee_id "
            "WHERE s.employee_id = ? AND (s.shift_type LIKE ? OR s.detail LIKE ?) ORDER BY s.date DESC LIMIT 50",
            (emp_id, f'%{q}%', f'%{q}%')
        ).fetchall()
    for r in rows:
        results.append({
            'type': 'schedule',
            'date': r['date'],
            'shift_type': r['shift_type'],
            'detail': r['detail'],
            'employee_id': r['employee_id'],
            'name': r['name'],
        })

    # 메모 검색
    if is_admin:
        mrows = db.execute(
            "SELECT m.*, u.name FROM memos m LEFT JOIN users u ON m.employee_id = u.employee_id "
            "WHERE m.content LIKE ? ORDER BY m.date DESC LIMIT 50",
            (f'%{q}%',)
        ).fetchall()
    else:
        mrows = db.execute(
            "SELECT m.*, u.name FROM memos m LEFT JOIN users u ON m.employee_id = u.employee_id "
            "WHERE m.employee_id = ? AND m.content LIKE ? ORDER BY m.date DESC LIMIT 50",
            (emp_id, f'%{q}%')
        ).fetchall()
    for r in mrows:
        results.append({
            'type': 'memo',
            'date': r['date'],
            'content': r['content'],
            'employee_id': r['employee_id'],
            'name': r['name'],
        })
    return jsonify(results)

# ── API: Users (Admin) ──────────────────────────────────────

@app.route('/api/users')
@login_required
def api_get_users():
    db = get_db()
    if session.get('role') == 'admin':
        rows = db.execute("SELECT employee_id, name, team, role FROM users ORDER BY team, name").fetchall()
    else:
        rows = db.execute(
            "SELECT employee_id, name, team FROM users WHERE employee_id = ?",
            (session['user_id'],)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/users', methods=['POST'])
@login_required
@admin_required
def api_add_user():
    data = request.json
    db = get_db()
    emp_id = str(data['employee_id']).strip()
    try:
        db.execute(
            "INSERT INTO users (employee_id, name, password, team, role) VALUES (?, ?, ?, ?, ?)",
            (emp_id, data['name'], hash_pw(emp_id), data.get('team', ''), 'user')
        )
        db.commit()
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': '이미 존재하는 사번입니다'}), 400

@app.route('/api/users/<emp_id>', methods=['DELETE'])
@login_required
@admin_required
def api_delete_user(emp_id):
    if emp_id == 'admin':
        return jsonify({'error': '관리자 계정은 삭제할 수 없습니다'}), 400
    db = get_db()
    db.execute("DELETE FROM users WHERE employee_id = ?", (emp_id,))
    db.execute("DELETE FROM schedules WHERE employee_id = ?", (emp_id,))
    db.execute("DELETE FROM memos WHERE employee_id = ?", (emp_id,))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/users/<emp_id>/reset-password', methods=['POST'])
@login_required
@admin_required
def api_reset_password(emp_id):
    db = get_db()
    db.execute("UPDATE users SET password = ? WHERE employee_id = ?",
               (hash_pw(emp_id), emp_id))
    db.commit()
    return jsonify({'success': True})

# ── Excel Import ────────────────────────────────────────────

SHIFT_MAP = {
    'D': 'D', 'S': 'S', 'G': 'G',
    '휴': '휴', 'O': 'O', '월중': '휴',
    '체': '체',
}

def parse_shift_cell(value):
    """셀 값에서 근무 타입과 상세 정보 분리"""
    if value is None:
        return '', ''
    val = str(value).strip()
    if not val:
        return '', ''
    lines = val.split('\n')
    shift = lines[0].strip()
    # 첫 줄이 근무타입인지 확인
    main_shift = ''
    for key in SHIFT_MAP:
        if shift.startswith(key):
            main_shift = SHIFT_MAP[key]
            break
    if not main_shift and shift in ('월중', '월중\n휴무'):
        main_shift = '휴'

    detail = '\n'.join(l.strip() for l in lines if l.strip()).replace(shift, '', 1).strip()
    if not main_shift:
        main_shift = shift
        detail = '\n'.join(lines[1:]).strip() if len(lines) > 1 else ''

    return main_shift, detail

@app.route('/api/import-excel', methods=['POST'])
@login_required
@admin_required
def api_import_excel():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Excel 파일만 업로드 가능합니다'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)

    try:
        result = import_excel(filepath)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def import_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    db = sqlite3.connect(DB_PATH)
    user_count = 0
    schedule_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Row 2: dates
        dates = []
        for col in range(5, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if isinstance(val, datetime):
                dates.append((col, val.strftime('%Y-%m-%d')))

        if not dates:
            continue

        # Row 5+: employees
        current_team = ''
        for row in range(5, ws.max_row + 1):
            team_cell = ws.cell(row=row, column=2).value
            if team_cell:
                current_team = str(team_cell).strip()

            name = ws.cell(row=row, column=3).value
            emp_id_raw = ws.cell(row=row, column=4).value
            if not name or not emp_id_raw:
                continue

            emp_id = str(emp_id_raw).strip().lstrip('0') if str(emp_id_raw).strip().isdigit() else str(emp_id_raw).strip()
            # 사번 원본 유지 (앞자리 0 포함)
            emp_id = str(emp_id_raw).strip()
            if isinstance(emp_id_raw, (int, float)):
                emp_id = str(int(emp_id_raw))
            name = str(name).strip()

            # 사용자 등록
            try:
                db.execute(
                    "INSERT OR IGNORE INTO users (employee_id, name, password, team, role) VALUES (?, ?, ?, ?, ?)",
                    (emp_id, name, hashlib.sha256(emp_id.encode()).hexdigest(), current_team, 'user')
                )
                user_count += 1
            except Exception:
                pass

            # 일정 등록
            for col, date_str in dates:
                cell_val = ws.cell(row=row, column=col).value
                shift_type, detail = parse_shift_cell(cell_val)
                if shift_type:
                    db.execute('''
                        INSERT INTO schedules (employee_id, date, shift_type, detail, updated_at)
                        VALUES (?, ?, ?, ?, datetime('now'))
                        ON CONFLICT(employee_id, date) DO UPDATE SET
                            shift_type=excluded.shift_type,
                            detail=excluded.detail,
                            updated_at=datetime('now')
                    ''', (emp_id, date_str, shift_type, detail))
                    schedule_count += 1

    db.commit()
    db.close()
    return {'success': True, 'users': user_count, 'schedules': schedule_count}

# ── iCal export (for Google/Naver calendar sync) ────────────

@app.route('/api/ical/<emp_id>.ics')
def api_ical_export(emp_id):
    """iCal 형식으로 일정 내보내기 - 구글/네이버 캘린더 연동용"""
    db_conn = sqlite3.connect(DB_PATH)
    db_conn.row_factory = sqlite3.Row
    user = db_conn.execute("SELECT * FROM users WHERE employee_id = ?", (emp_id,)).fetchone()
    if not user:
        db_conn.close()
        return "Not found", 404

    schedules = db_conn.execute(
        "SELECT * FROM schedules WHERE employee_id = ? ORDER BY date", (emp_id,)
    ).fetchall()
    memos = db_conn.execute(
        "SELECT * FROM memos WHERE employee_id = ? ORDER BY date", (emp_id,)
    ).fetchall()
    db_conn.close()

    shift_names = {
        'D': '주간근무(D)', 'S': '석간근무(S)', 'G': '야간근무(G)',
        '휴': '휴무', 'O': '결근/기타', '체': '체력단련',
        '월중': '월중휴무',
    }

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//ScheduleApp//KR",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:{user['name']} 근무일정",
    ]

    for s in schedules:
        st = s['shift_type']
        summary = shift_names.get(st, st)
        desc = s['detail'].replace('\n', '\\n') if s['detail'] else ''
        date_str = s['date'].replace('-', '')
        uid = f"{s['employee_id']}-{s['date']}@schedule-app"
        lines.extend([
            "BEGIN:VEVENT",
            f"DTSTART;VALUE=DATE:{date_str}",
            f"DTEND;VALUE=DATE:{date_str}",
            f"SUMMARY:{summary}",
            f"DESCRIPTION:{desc}" if desc else f"DESCRIPTION:{summary}",
            f"UID:{uid}",
            "END:VEVENT",
        ])

    for m in memos:
        date_str = m['date'].replace('-', '')
        uid = f"memo-{m['id']}@schedule-app"
        lines.extend([
            "BEGIN:VEVENT",
            f"DTSTART;VALUE=DATE:{date_str}",
            f"DTEND;VALUE=DATE:{date_str}",
            f"SUMMARY:[메모] {m['content'][:30]}",
            f"DESCRIPTION:{m['content'].replace(chr(10), '\\n')}",
            f"UID:{uid}",
            "END:VEVENT",
        ])

    lines.append("END:VCALENDAR")
    ical_content = '\r\n'.join(lines)

    from flask import Response
    return Response(
        ical_content,
        mimetype='text/calendar',
        headers={'Content-Disposition': f'attachment; filename={emp_id}.ics'}
    )

# ── PWA ─────────────────────────────────────────────────────

@app.route('/manifest.json')
def manifest():
    return jsonify({
        "name": "일정관리",
        "short_name": "일정",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#4A90D9",
        "icons": [
            {"src": "/static/icons/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icons/icon-512.png", "sizes": "512x512", "type": "image/png"}
        ]
    })

@app.route('/sw.js')
def service_worker():
    return app.send_static_file('js/sw.js'), 200, {'Content-Type': 'application/javascript'}

# ── Session info ────────────────────────────────────────────

@app.route('/api/session')
@login_required
def api_session():
    return jsonify({
        'employee_id': session['user_id'],
        'name': session['user_name'],
        'role': session.get('role', 'user'),
    })

# ── Auto-import default Excel on first run ──────────────────

def auto_import_default():
    """기본 엑셀 파일 자동 임포트"""
    try:
        db = sqlite3.connect(DB_PATH)
        count = db.execute("SELECT COUNT(*) FROM schedules").fetchone()[0]
        db.close()
        if count == 0:
            default_path = os.path.join(os.path.dirname(__file__), '..', '2026년_일정.xlsx')
            if os.path.exists(default_path):
                result = import_excel(default_path)
                print(f"[자동 임포트] 사용자 {result['users']}명, 일정 {result['schedules']}건 등록")
            else:
                print("[자동 임포트] 엑셀 파일 없음 - 건너뜀")
    except Exception as e:
        print(f"[자동 임포트 실패] {e}")

# ── Health check ─────────────────────────────────────────────

@app.route('/healthz')
def healthz():
    return 'ok', 200

# ── Init ─────────────────────────────────────────────────────

with app.app_context():
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    init_db()
    auto_import_default()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
